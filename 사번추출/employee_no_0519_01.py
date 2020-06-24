import os
from openpyxl import Workbook
from openpyxl import load_workbook
import pyautogui
from tkinter import Tk
from datetime import datetime

def copy_right_value(file_name, tab_num, route):
    image_search_click(file_name, route)
    # 탭
    for i in range(tab_num):
        pyautogui.press('tab')
    pyautogui.PAUSE = 0.1
    # 사번 클립보드에 복사
    pyautogui.hotkey('ctrl', 'c')
    pyautogui.PAUSE = 0.1
    copied_value = Tk().clipboard_get()
    return copied_value

def save_value_in_excel(ws, current_copied_employee_id, a, column_num, route):
    os.chdir(route)
    # 대상 셀 객체화
    pyautogui.PAUSE = 0.1
    target_cell = ws.cell(a+1,column_num)
    target_cell.value = current_copied_employee_id

def image_search_click(file_name, route):
    # ※주의※ 파일명 영어만 됨
    os.chdir(route)
    while True:
        location = pyautogui.locateCenterOnScreen(file_name)
        if location == None:
            print(file_name+'이미지 찾지 못함')
            continue
        else:
            pyautogui.doubleClick(location)
            pyautogui.PAUSE = 0.1
            break

def image_search_only(file_name, route):
    # ※주의※ 파일명 영어만 됨
    os.chdir(route)
    while True:
        location = pyautogui.locateCenterOnScreen(file_name)
        if location == None:
            print(file_name+'이미지 찾지 못함')
            continue
        else:
            print(file_name+'이미지 찾음')
            pyautogui.PAUSE = 0.1
            break

def down_check(a, max_row_num):
    mod = a % max_row_num
    if mod != 0:
        pyautogui.press('down')

def file_save(origin_folder, file_subject):
    os.chdir(origin_folder)
    wb.save(file_subject)

def page_change_check(a, max_row_num, image_folder):
    mod = a % max_row_num
    os.chdir(image_folder)

    if mod == 0 and a != 0:
        image_search_click('page_change.png', image_folder)
        pyautogui.PAUSE = 5
        while True:
            temp_employee_id = copy_right_value('employee_id.png', tab_num = 1, route = image_folder)
            image_search_click('status.png', image_folder)
            if len(temp_employee_id) == 5:
                break



# 메인 코드
origin_folder = 'D:\\세환\\python\\사번추출'
image_folder = 'D:\\세환\\python\\사번추출\\image'
file_subject = 'employee_no_result' + '.xlsx'
file_matter = 'N'
previous_copy_id = 'None'
len_of_cpdid = 'init'

if file_subject in os.listdir(origin_folder):
    os.chdir(origin_folder)
    file_matter = 'Y'
    wb = load_workbook(file_subject)
    ws = wb.active
    init_cell_row = ws.max_row

else:
    wb = Workbook()
    ws = wb.active


if file_matter == 'Y':
    row_num = init_cell_row
else:
    row_num = 0
total_no = 1000

for a in range(total_no):
    # 활성화 줄 클릭
    image_search_click('active_list_value.png', route=image_folder)
    image_search_only('active_flag.png', route=image_folder)
    if a != 0:
        pyautogui.press('down')
    pyautogui.PAUSE = 0.1
    while True:
        current_copied_employee_id = copy_right_value('employee_id.png', tab_num = 1, route = image_folder)
        len_of_cpdid = len(current_copied_employee_id)
        if current_copied_employee_id != previous_copy_id and len_of_cpdid == 5:
            break
    previous_copy_id = current_copied_employee_id
    len_of_cpdid = 'init'
    save_value_in_excel(ws, current_copied_employee_id, a+row_num, column_num = 1, route = origin_folder)
    pyautogui.PAUSE = 0.1
    # 파일 저장기능
    file_save(origin_folder=origin_folder, file_subject = file_subject)
