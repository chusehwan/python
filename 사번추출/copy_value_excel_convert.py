import os
from openpyxl import Workbook
from openpyxl import load_workbook
import pyautogui
from tkinter import Tk
from datetime import datetime
import pyperclip



def value_make(copied_value):
    divide_values = copied_value.split()
    return divide_values

def file_save(origin_folder, file_subject):
    os.chdir(origin_folder)
    wb.save(file_subject)

origin_folder = 'D:\\세환\\python\\사번추출'
file_subject = 'employee_no_result' + '.xlsx'
os.chdir(origin_folder)

wb = Workbook()
ws = wb.active
copy_v = pyperclip.paste()
divided_values = value_make(copied_value=copy_v)
list_len = len(divided_values)

for i in range(list_len):
    ws.cell(i+1,1).value = divided_values[i]

file_save(origin_folder=origin_folder, file_subject = file_subject)



#position = value_make(copied_value=copy_v, column = '직위')
#position2 = value_make(copied_value=copy_v, column = '직책')

origin_folder = 'D:\\세환\\python\\사번추출'
image_folder = 'D:\\세환\\python\\사번추출\\image'
file_subject = 'employee_no_result' + '.xlsx'

#print(os.listdir(origin_folder))
#if file_subject in os.listdir(origin_folder):
 #   print('파일 있음')







