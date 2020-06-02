import os
from openpyxl import Workbook
from openpyxl import load_workbook
import pyautogui
from tkinter import Tk
from datetime import datetime
import pyperclip

class divide_list:
    def __init__(self, original_list):
        self.original_list = original_list

    def get_hierachy_data(self):
        divided_by_n_list = self.original_list.split('\n')
        iter = len(divided_by_n_list)
        for i in range(iter):
            if divided_by_n_list[i] == '기본정보 ':
                return (divided_by_n_list[i-2])

    def get_name(self):
        divided_by_n_list = self.original_list.split('\n')
        iter = len(divided_by_n_list)
        for i in range(iter):
            if divided_by_n_list[i] == '기본정보 ':
                return (divided_by_n_list[i+1])

    def get_account(self):
        divided_by_n_list = self.original_list.split('\n')
        iter = len(divided_by_n_list)
        for i in range(iter):
            temp = divided_by_n_list[i]
            if '회계코드' in temp:
                return temp

    def get_position(self):
        divided_by_n_list = self.original_list.split('\n')
        iter = len(divided_by_n_list)
        for i in range(iter):
            temp = divided_by_n_list[i]
            if '직위' in temp:
                position =  temp
                break
        try :
            position_list = position.split()
            for i in range(len(position_list)):
                if position_list[i] == '직위':
                    return position_list[i+1]
        except UnboundLocalError:
            pass

    def get_position2(self):
        divided_by_n_list = self.original_list.split('\n')
        iter = len(divided_by_n_list)
        for i in range(iter):
            temp = divided_by_n_list[i]
            if '직책' in temp:
                position =  temp
                break
        try :
            position_list = position.split()
            for i in range(len(position_list)):
                if position_list[i] == '직책':
                    return position_list[i+1]
        except UnboundLocalError:
            pass

    def get_email(self):
        divided_by_n_list = self.original_list.split('\n')
        iter = len(divided_by_n_list)
        for i in range(iter):
            temp = divided_by_n_list[i]
            if '팩스번호' in temp:
                position =  temp

        temp = position.split()
        ra = len(temp)
        for i in range(ra):
            if temp[i] == 'E-Mail':
                return temp[i+1]

def value_make(copied_value):
    divide_values = copied_value.split('나누기')
    return divide_values

def file_save(origin_folder, file_subject):
    os.chdir(origin_folder)
    wb.save(file_subject)

origin_folder = 'D:\\세환\\python\\사번추출'
file_subject = 'excel_convert' + '.xlsx'
os.chdir(origin_folder)

wb = Workbook()
ws = wb.active
copy_v = pyperclip.paste()
divided_values = value_make(copied_value=copy_v)
row_num = 1

for i in divided_values:
    try:
        person_info = divide_list(i)
        name = person_info.get_name()
        account = person_info.get_account()
        position = person_info.get_position()
        position2 = person_info.get_position2()
        Email_num = person_info.get_email()
        hierachy = person_info.get_hierachy_data()
    except IndexError:
        pass
    ws.cell(row_num, 3).value = name
    ws.cell(row_num, 4).value = position
    ws.cell(row_num, 5).value = position2
    ws.cell(row_num, 6).value = account
    ws.cell(row_num, 7).value = Email_num
    ws.cell(row_num, 8).value = hierachy

    row_num += 1
    print(row_num)
file_save(origin_folder=origin_folder, file_subject=file_subject)
print('완료')






