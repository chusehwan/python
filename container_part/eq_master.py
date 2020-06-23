from openpyxl import Workbook
from tkinter import Tk
import functions as fn
import switch_app as sa
import pyautogui
import time
from os import chdir
import pyperclip

image_folder = r'D:\세환\python\container_part\eq_master'
save_folder = 'D:\\'
title  = 'eq_master_'
# Work Book열어서 최종결과물 파일 만들기
wb = Workbook()
ws = wb.active
# 클립보드에 있는 값 CNTR No, POL DT, POD DT로 분리하고 엑셀에 저장
# 전체 컨테이너 넘버 수량 파악
ws, num_of_cntr, container_nums = fn.get_value_make_excel_for_dp(ws)
file_subject = fn.get_subject(title)

# 익스플로러 창 열기
sa.switch_app()
fn.sleep(1)

for i in range(num_of_cntr):

    # EQ NO입력란 클릭 및 기입력값 지우기
    while True:
        try:
            fn.image_search_double_click('eq_no.png', image_folder, x=100, confidence_parameter = 0.7)
            break
        except :
            continue
    # 컨테이너값 입력하고 서치(=F2)
    fn.data_clear_input(container_nums[i])
    fn.sleep(1)
    pyautogui.press('f2')
    fn.sleep(0.1)
    pyautogui.press('enter')
    fn.sleep(0.1)
    # 결과 나왔는지 여부 확인
    fn.image_search_only('result_confirm.png', image_folder, confidence_parameter = 0.8)

    # Ship Hold Y/N값 확인
    ship_hold_yn = ''
    pyperclip.copy('')
    fn.image_search_double_click('ship_hold_yn.png', image_folder, x=60, confidence_parameter = 0.9)
    pyautogui.hotkey('ctrl','c')
    fn.sleep(0.1)
    try:
        ship_hold_yn = Tk().clipboard_get()
    except:
        pass
    # 'N' 일경우 continue, Yes일경우 클릭하여 No로 바꿔줌
    if ship_hold_yn == 'N' or ship_hold_yn == ' ':
        check_hold = 'N'

        ws.cell(i + 2, 2).value = ship_hold_yn
        ws.cell(i + 2, 1).value = container_nums[i]
        continue

    while True:
        # 쉽 홀드 yn이 y인경우 n로 바꿔주고 바뀌었는지 다시 클릭하여 확인하기
        if ship_hold_yn == 'Y' or ship_hold_yn == '':
            for a in range(3):
                pyautogui.press('down')
            fn.image_search_double_click('space_for_click.png', image_folder, confidence_parameter = 0.9)
            # Ship Hold Y/N값 확인
            ship_hold_yn = ''
            fn.image_search_double_click('ship_hold_yn.png', image_folder, x=60, confidence_parameter=0.9)
            pyautogui.hotkey('ctrl', 'c')
            fn.sleep(0.1)
            ship_hold_yn = Tk().clipboard_get()
        if ship_hold_yn == 'N':
            break
    pyautogui.press('tab')
    fn.sleep(0.1)
    # 페이지 다운하기 위한 ship hold 클릭 및 page down 후 save
    while True:
        while True:
            try:
                save_result = fn.image_search_only_once('save_button.png', image_folder)
                break
            except:
                continue
        if save_result == 'find_image':
            fn.image_search_double_click('save_button.png', image_folder, confidence_parameter=0.9)
            break
        else:
            fn.image_search_double_click('space_for_click.png', image_folder, confidence_parameter=0.9)
            fn.sleep(0.1)
            pyautogui.press('pgdn')
            fn.sleep(0.1)
    for a in range(3):
        pyautogui.press('enter')
        fn.sleep(0.1)
    # Remark란에 저장완료 표시
    ws.cell(i + 2, 2).value = ship_hold_yn
    ws.cell(i+2, 1).value = container_nums[i]
    chdir(save_folder)
    wb.save(file_subject)
    # 오른쪽 상단 기입력된 CNTR넘버 Delete하기

print('프로그램 끝 ^0^')
print('총소요시간(초) : ', round(time.perf_counter(), 2))
print('건당소요시간(초) : ', round((time.perf_counter()/num_of_cntr), 2))