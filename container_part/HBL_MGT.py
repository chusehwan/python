from openpyxl import Workbook
from tkinter import Tk
import functions as fn
import switch_app as sa
import pyautogui
import time
from os import chdir
import pyperclip

image_folder = r'D:\세환\python\container_part\image_HBL_MGT'
save_folder = 'D:\\'
title  = 'HBL_MGT_'
CARRIER ='1128722'
VESSEL = 'transoptima'
VOY = 'trans'
AGENT = '1128722'
OPERATOR = '70477'

# Work Book열어서 최종결과물 파일 만들기
wb = Workbook()
ws = wb.active
# 클립보드에 있는 값 CNTR No, POL DT, POD DT로 분리하고 엑셀에 저장
# 전체 컨테이너 넘버 수량 파악
ws, num_of_so, so_nums = fn.get_value_make_excel_for_so(ws)
file_subject = fn.get_subject(title)

# 익스플로러 창 열기
sa.switch_app()
fn.sleep(1)

for i in range(num_of_so):
    # SO No 클릭 후 지우기, so_nums[i] 값 입력, F2통하여 조회하기
    fn.image_search_click('For_SO_No_Click.png', image_folder, x=60, y=-5, confidence_parameter = 0.8)
    pyautogui.hotkey('ctrl', 'a')
    fn.sleep(0.1)
    fn.data_clear_input(so_nums[i])
    fn.sleep(0.1)
    pyautogui.press('f2')
    fn.sleep(0.1)
    # 값나올때까지 대기(오른쪽의 so넘버 더블클릭 및 복사 후 so_nums[i]와 동일한지 비교후 같으면 loop벗어나기
    while True:
        pyperclip.copy('')
        now_so_no = None
        fn.image_search_click('confirm_so_no.png', image_folder)
        fn.image_search_double_click('confirm_so_no.png', image_folder, x= 28, y=-15)
        pyautogui.hotkey('ctrl', 'c')
        fn.sleep(0.1)
        try:
            now_so_no = Tk().clipboard_get()
        except:
            continue
        if now_so_no is not None:
            now_so_no = now_so_no.strip()
        if now_so_no == so_nums[i]:
            break
    pyperclip.copy('')
    # 'BL'부분 클릭 후 오른쪽에 BL No가 나오는지 여부 검사
    fn.image_search_click('BL_No.png', image_folder, x=200)
    fn.sleep(0.1)
    pyautogui.hotkey('ctrl', 'a')
    fn.sleep(0.1)
    pyautogui.hotkey('ctrl', 'c')
    # BL넘버가 이미 생성되어있는지 판단 후 없으면 넘어가고 있으면 다음 SO 조회로 가기
    try:
        dbl_no = Tk().clipboard_get()
        fn.sleep(0.1)
        ws.cell(i + 2, 1).value = so_nums[i]
        ws.cell(i + 2, 2).value = dbl_no
        chdir(save_folder)
        wb.save(file_subject)
        continue
    except:
        pass
    dbl_no = None
    fn.image_search_click('DBL.png', image_folder)
    # 없을경우 각 부분클릭하여 값입력
    # 1)Carrier 부분에 '1128722', 2) VSL에 'transoptima', 3)VOY에 'trans', (4) opeartor에 70477 입력
    fn.image_search_double_click('Carrier.png', image_folder, x= 40)
    fn.sleep(0.1)
    pyautogui.write(CARRIER)
    fn.sleep(0.1)
    pyautogui.press('tab')
    fn.sleep(0.1)
    fn.image_search_click('Confirm.png', image_folder)
    fn.sleep(0.1)
    for a in range(3):
        pyautogui.press('tab')
        fn.sleep(0.1)
    pyautogui.write(VESSEL)
    fn.sleep(0.1)
    pyautogui.press('tab')
    fn.sleep(0.1)
    pyautogui.write(VOY)
    fn.sleep(0.1)
    fn.image_search_double_click('Operator.png', image_folder, x=40)
    fn.sleep(0.1)
    pyautogui.write(OPERATOR)
    fn.sleep(0.1)
    pyautogui.press('tab')
    fn.sleep(0.1)
    # Save버튼 누르고 저장되었는지 여부 확인
    fn.image_search_click('save.png', image_folder)
    err_check = ''
    close_check = ''
    while True:
        close_check = fn.image_search_only_once('before_confirm.png', image_folder)
        err_check = fn.image_search_only_once('Equipment_err.png', image_folder)
        if close_check == 'find_image':
            break
        if err_check == 'find_image':
            break
    if err_check == 'find_image':
        ws.cell(i + 2, 1).value = so_nums[i]
        ws.cell(i + 2, 2).value = 'Error'
        continue
    fn.image_search_click('before_confirm.png', image_folder)
    pyautogui.press('enter')
    fn.sleep(0.1)
    fn.image_search_click('close.png', image_folder)
    fn.sleep(0.1)
    # 저장되면 'DB'부분 BL넘버 더블클릭 및 복사, 복사되었는지 확인
    pyperclip.copy('')
    dbl_no = None
    while True:
        fn.image_search_double_click('DBL.png', image_folder)
        fn.sleep(0.1)
        pyautogui.press('tab')
        fn.sleep(0.1)
        pyautogui.hotkey('ctrl', 'a')
        fn.sleep(0.1)
        pyautogui.hotkey('ctrl', 'c')
        # BL넘버가 이미 생성되어있는지 판단 후 없으면 넘어가고 있으면 다음 SO 조회로 가기
        try:
            dbl_no = Tk().clipboard_get()
        except:
            continue
        fn.sleep(0.1)
        break
    # 저장된 BL번호에 대해 ws.cell(i + 2, 3).value = 에 복사된 BL값 기입
    ws.cell(i+2, 1).value = so_nums[i]
    ws.cell(i+2, 2).value = dbl_no
    chdir(save_folder)
    wb.save(file_subject)

print('프로그램 끝 ^0^')
print('총소요시간(초) : ', round(time.perf_counter(), 2))
print('건당소요시간(초) : ', round((time.perf_counter()/num_of_so), 2))