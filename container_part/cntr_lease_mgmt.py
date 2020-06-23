from openpyxl import Workbook
from tkinter import Tk
import functions as fn
import switch_app as sa
import pyautogui
import time
from os import chdir
import pyperclip

image_folder = r'D:\세환\python\container_part\image_cntr_lease_mgmt'
save_folder = 'D:\\'
title  = 'CNTR_Lease_MGMT_'
# Work Book열어서 최종결과물 파일 만들기(1열 so정보, 2열 cntr no, 3열 POL DT, 4열 POD DT)
wb = Workbook()
ws = wb.active
# 클립보드에 있는 값 CNTR No, POL DT, POD DT로 분리하고 엑셀에 저장
# 전체 컨테이너 넘버 수량 파악
ws, num_of_cntr = fn.get_value_make_excel_2(ws)
file_subject = fn.get_subject(title)

# 익스플로러 창 열기
sa.switch_app()
fn.sleep(1)
# 오른쪽 상단 No의 contract no 부분 클릭하여 상태를 CNTR No(Appointed)로 바꾸기
fn.image_search_click('cntr_no_appointed_select.png', image_folder, x=-30, y=20)
fn.sleep(0.1)
pyautogui.press('pgdn')
fn.sleep(0.1)

for i in range(num_of_cntr):
    # 컨테이너 입력부분 값 삭제
    fn.image_search_double_click('cntr_no_appointed_delete.png', image_folder, y=10)
    # 오른쪽 상단 No의 하단 하얀색 부분 클릭하여 컨넘버 붙여넣기 후 F2로 조회
    fn.image_search_double_click('cntr_no_appointed_select.png', image_folder, y=100)
    cntr_no = ws.cell(i+2,1).value
    pyperclip.copy(cntr_no)
    fn.sleep(0.1)
    pyautogui.hotkey('ctrl', 'v')
    fn.sleep(0.1)
    pyautogui.press('f2')
    # 결과물 나왔는지 확인
    fn.image_search_only('check_result.png', image_folder)

    # 왼쪽하단 Return LOC 의 Code값이 있는지 검사
    return_loc_value = fn.check_column_value('Return_LOC.png', image_folder,x=-40, y=40)
    # 없으면 아래절차 진행, 있으면 위로 돌아가기
    if return_loc_value != '':
        ws.cell(i+2, 2).value = '값있음'
        continue
    while True:
        while True:
            try:
                # 왼쪽하단 Lease LOC 의 Code부분 더블클릭하여 복사
                Lease_loc_value = fn.check_column_value('Lease_LOC.png', image_folder, x=-40, y=40)
                fn.sleep(0.1)
                # tab하여 Lease DT복사
                pyautogui.press('tab')
                fn.sleep(0.1)
                pyautogui.press('enter')
                fn.sleep(0.1)
                pyautogui.hotkey('ctrl','c')
                fn.sleep(0.1)
                Lease_dt_value = Tk().clipboard_get()
                one_day_added_lease_dt_value = fn.convert_date(Lease_dt_value, add_day = 1)
                break
            except:
                continue
        # tab하고 Enter후 Return LOC에 앞에서 복사된 Lease_LOC 값 붙여넣기
        fn.sleep(0.1)
        pyautogui.press('tab')
        fn.sleep(0.1)
        pyautogui.press('enter')
        fn.sleep(0.1)
        pyautogui.write(Lease_loc_value)
        fn.sleep(0.1)
        # tab하고 Enter후 Return DT에 앞에서 복사된 Lease_DT 값 붙여넣기
        pyautogui.press('tab')
        fn.sleep(0.1)
        pyautogui.press('enter')
        fn.sleep(0.1)
        pyautogui.write(one_day_added_lease_dt_value)
        fn.sleep(0.1)
        return_loc_value = fn.check_column_value('Return_LOC.png', image_folder,x=-40, y=40)
        # 없으면 아래절차 진행, 있으면 위로 돌아가기
        if return_loc_value != '':
            break
    # SAVE버튼 클릭
    fn.image_search_double_click('save_button.png', image_folder, x=30, y=20, confidence_parameter = 0.9)
    fn.sleep(0.3)
    for a in range(3):
        pyautogui.press('enter')
        fn.sleep(0.1)
    # Remark란에 저장완료 표시
    ws.cell(i + 2, 2).value = 'ok'
    chdir(save_folder)
    wb.save(file_subject)
    # 오른쪽 상단 기입력된 CNTR넘버 Delete하기

print('프로그램 끝 ^0^')
print('총소요시간(초) : ', round(time.perf_counter(), 2))
print('건당소요시간(초) : ', round((time.perf_counter()/num_of_cntr), 2))