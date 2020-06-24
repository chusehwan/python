from openpyxl import Workbook
from tkinter import Tk
import functions as fn
import switch_app as sa
import pyautogui
import time
from os import chdir

image_folder = r'D:\세환\python\container_part\image_so_creation'
save_folder = 'D:\\'
original_so_no = 'SHQF200600006463'
operator = '70477'
title  = 'SO_Reg_result_'
# Work Book열어서 최종결과물 파일 만들기(1열 so정보, 2열 cntr no, 3열 POL DT, 4열 POD DT)
wb = Workbook()
ws = wb.active
# 클립보드에 있는 값 CNTR No, POL DT, POD DT로 분리하고 엑셀에 저장
# 전체 컨테이너 넘버 수량 파악
ws, num_of_cntr = fn.get_value_make_excel(ws)
file_subject = fn.get_subject(title)

# 익스플로러 창 열기
sa.switch_app()
fn.sleep(1)

for i in range(num_of_cntr):
    # # work sheet에 저장되어있는 cntr , dt정보 가져오기
    cntr_no = ws.cell(i+2, 2).value
    pol_dt = ws.cell(i+2, 3).value
    pod_dt = ws.cell(i+2, 4).value
    # SO No 입력창 클릭
    fn.image_search_double_click('SO_No.png', image_folder, x=80)
    # 최초 원본 SO No 입력(SHQF200600006463)
    fn.data_clear_input(original_so_no)
    fn.sleep(1)
    pyautogui.press('f2')
    # Service Step Tab 클릭
    fn.both_image_search_double_click(file_name_1='Service_1.png',file_name_2='Service_2.png', image_folder = image_folder)
    # SO정보 로드 되었는지 확인
    fn.image_search_only('so_load_complete.png', image_folder)
    # so copy하기
    fn.sleep(2)
    fn.image_search_double_click('Copy_Button.png', image_folder, x=-20, confidence_parameter = 0.9)
    # copy되었는지 확인
    fn.image_check_n_click('copy_check.png', 'Copy_Button.png', image_folder)
    time.sleep(0.3)
    # cntr no 앞 3자리가 PKEU 면 Shipment Type 을 클릭하여 'Match Back'선택하기 아니면 pass
    fn.check_PKEU(cntr_no, image_folder)
    # Port of Loading 의 Order Plan에 POL DT값 입력
    fn.input_dt(pol_dt, 'Port_of_Loading.png', image_folder)
    # 에러메시지(past 데이트 입력 불가) 발생 시 close버튼 클릭
    fn.check_date_err_click_close('Close_Button.png', port_date = pol_dt, image_folder=image_folder)
    # Place of Delivery 의 Order Plan에 POD DT값 입력
    fn.input_dt(pod_dt, 'Place_of_Delivery.png', image_folder)
    # 에러메시지(past 데이트 입력 불가) 발생 시 close버튼 클릭
    fn.check_date_err_click_close('Close_Button.png', port_date = pod_dt, image_folder=image_folder)
    # operator 바꾸기
    fn.image_search_double_click('Operator_DEPT.png', image_folder)
    pyautogui.hotkey('ctrl', 'a')
    fn.sleep(0.1)
    pyautogui.press('backspace')
    fn.sleep(0.1)
    pyautogui.press('tab')
    fn.sleep(0.1)
    fn.image_search_double_click('Operator_DEPT.png', image_folder)
    fn.sleep(0.1)
    pyautogui.write(operator)
    fn.sleep(0.1)
    pyautogui.press('tab')
    # Shipment Tab 클릭
    fn.image_search_click('Shipment_Tab.png', image_folder)
    # 왼쪽하단 Container Summary에 40FT HC 입력, CNTR No 앞자리가 PKEU면 SOC, 아니며 DP클릭 1대 입력
    fn.image_search_click('Summary_Add_Button.png', image_folder)
    fn.image_search_click('Summary_Size.png', image_folder,x=0, y=30)
    pyautogui.write('4')
    fn.sleep(0.1)
    pyautogui.press('down')
    fn.sleep(0.1)
    pyautogui.press('enter')
    fn.sleep(0.1)
    pyautogui.press('tab')
    fn.sleep(0.1)
    pyautogui.write('H')
    fn.sleep(0.1)
    pyautogui.press('down')
    fn.sleep(0.1)
    pyautogui.press('enter')
    fn.sleep(0.1)
    pyautogui.press('tab')
    fn.sleep(0.1)
    if cntr_no[:4] == 'PKEU':
        pyautogui.write('S')
        fn.sleep(0.1)
    else:
        pyautogui.write('D')
        fn.sleep(0.1)
    pyautogui.press('tab')
    fn.sleep(0.1)
    pyautogui.write('1')
    fn.sleep(0.1)
    pyautogui.press('enter')
    fn.sleep(0.1)
    # Detail 란에 Container No입력, GWT 2000입력, Volume 1 입력, PKG 1 입력
    fn.image_search_click('Detail_Container_No.png', image_folder, x=0, y=30)
    pyautogui.write(cntr_no)
    fn.sleep(0.1)
    pyautogui.press('tab')
    fn.sleep(0.1)
    pyautogui.press('tab')
    fn.sleep(0.1)
    # G.WT입력
    pyautogui.write('2200')
    fn.sleep(0.1)
    pyautogui.press('tab')
    fn.sleep(0.1)
    # Volume 입력
    pyautogui.write('1')
    # PKG 입력
    fn.sleep(0.1)
    pyautogui.press('tab')
    fn.sleep(0.1)
    pyautogui.write('1')
    fn.sleep(0.1)
    # Service Step Tab으로 이동하여 PTD 에다가 POL DT + 0000 입력, PTA 에다가 POD DT +0000 입력
    fn.image_search_double_click('Service_Step_Tab.png', image_folder)
    # 1구간 EO가 있는 경우 삭제해주기
    fn.check_err_click_close('stage_1_check.png', 'Service_Step_Delete_Button.png', image_folder)
    fn.image_search_only('PTD.png', image_folder)
    fn.image_search_click('PTD.png', image_folder, y=30)
    fn.sleep(0.1)
    pyautogui.write(pol_dt+'0000')
    fn.sleep(0.1)
    fn.image_search_click('PTA.png', image_folder, y=30)
    fn.sleep(0.1)
    pyautogui.write(pod_dt+'0000')
    fn.sleep(0.1)
    pyautogui.press('tab')
    # SAVE
    fn.image_search_click('Save.png', image_folder)
    # pol, pod date가 past일 경우 뜨는 팝업창 클릭
    fn.check_date_err_click_close_last('last_confirm_1.png', port_date_1=pol_dt, port_date_2=pod_dt, image_folder=image_folder)
    # 저장 진행과정 버튼 클릭
    fn.image_search_click('please_check_the_route.png', image_folder)
    fn.image_search_click('last_confirm_2.png', image_folder)
    fn.image_search_click('before_confirm.png', image_folder)
    fn.image_search_click('confirm.png', image_folder)
    # 에러메시지가 나오는지 save success나오는지 검토하여 최종 클릭
    last_result = None
    last_result = fn.last_check('err_pop_up.png', 'success_confrim_1.png', image_folder)
    pyautogui.press('enter')
    # SO생성 여부 확인하여 생성되면 Work Sheet상 so넘버를 기입하고 아니면 '생성안됨'표기
    if last_result == 'suc':
        while True:
            copied_SO_No = None
            fn.image_search_double_click('SO_No.png', image_folder)
            fn.sleep(0.1)
            pyautogui.hotkey('ctrl', 'a')
            fn.sleep(0.1)
            pyautogui.hotkey('ctrl','c')
            fn.sleep(0.1)
            copied_SO_No = Tk().clipboard_get()
            print(copied_SO_No+':'+cntr_no)
            if copied_SO_No:
                ws.cell(i + 2, 1).value = str(copied_SO_No)
                break
    if last_result == 'err':
        ws.cell(i+2, 1).value='생성안됨'
    fn.sleep(0.1)
    chdir(save_folder)
    wb.save(file_subject)

chdir(save_folder)
fn.sleep(0.1)
wb.save(file_subject)

print('프로그램 끝 ^0^')
print('총소요시간(초) : ', round(time.perf_counter(), 2))
print('건당소요시간(초) : ', round((time.perf_counter()/num_of_cntr), 2))