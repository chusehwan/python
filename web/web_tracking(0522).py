import os
from openpyxl import Workbook
import pyperclip
from datetime import datetime
from datetime import date
from selenium import webdriver
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

def get_subject():
    '''파일제목 마지막에 저장할 현재일시 및 파일제목 설정'''
    now = datetime.now()
    today = date.today()
    current_time = now.strftime("%H-%M-%S")
    file_name = 'Tracking_result_'+str(today)+'_'+str(current_time)+'.xlsx'
    return file_name

def value_make(copied_value):
    '''클립보드에서 가져온 값 나눠주기(앞3자리 뺀값으로 list에 저장)'''
    divide_values = copied_value.split()
    return divide_values

def verifi_awb(awb_list):
    '''awb넘버 바탕 앞 3자리 기준으로 대한항공,아시아나 정리'''
    ke_list=[]
    oz_list=[]
    for i in awb_list:
        if i[:3] == '180':
            ke_list.append(i)
        elif i[:3] == '988':
            oz_list.append(i)
    total_list = ke_list+oz_list
    return total_list, ke_list, oz_list

def file_save(origin_folder, file_subject):
    '''파일 지정 폴더에 저장'''
    os.chdir(origin_folder)
    wb.save(file_subject)

def get_tracking_info_ke(list, driver_dir, row_num):
    '''대한항공 홈페이지 접속 및 트랙킹 작업 awb list 바탕 진행'''
    ke_len = len(list)
    chromedriver_dir = driver_dir
    driver = webdriver.Chrome(chromedriver_dir)
    driver.get('https://cargo.koreanair.com/ko/tracking')
    wait = WebDriverWait(driver, 10)
    row_num = row_num
    for awb_num in list:
        # 브라우저가 나올때까지 대기 및 검색버튼 나오는지 확인 후 다음스탭 진행
        wait.until(EC.element_to_be_clickable((By.ID, 'awbDocNo1')))
        driver.find_element_by_xpath('//*[@id="awbDocNo1"]').send_keys(awb_num[3:])
        track_button = driver.find_element_by_xpath('//*[@id="tracking"]/div/div/div[1]/form/div[2]/div/div/div/div[2]/button')
        track_button.click()
        # 수량/중량 text나오는지 검사
        wait.until(EC.element_to_be_clickable((By.ID, 'tabClick')))
        try:
            pol_atd = driver.find_element_by_xpath('//*[@id="myTabContent"]/div/div[2]/div/div/div[2]/div[3]')
            pol_atd = convert_ke_time_info(pol_atd.text)
            ws.cell(row_num, 2).value = pol_atd
        except:
            ws.cell(row_num, 2).value = 'No_data'
        try:
            pod_ata = driver.find_element_by_xpath('//*[@id="myTabContent"]/div/div[4]/div/div/div[2]/div[3]')
            pod_ata = convert_ke_time_info(pod_ata.text)
            ws.cell(row_num, 3).value = pod_ata
        except:
            ws.cell(row_num, 3).value = 'No_data'
        driver.find_element_by_xpath('//*[@id="awbDocNo1"]').clear()
        time.sleep(0.1)
        row_num += 1
    print('대한항공 tracking 완료')
    driver.quit()
    return row_num

def get_tracking_info_oz(list, driver_dir, row_num):
    '''아시아나 홈페이지 접속 및 트랙킹 작업 awb list 바탕 진행'''
    oz_len = len(list)
    chromedriver_dir = driver_dir
    driver = webdriver.Chrome(chromedriver_dir)
    driver.get('https://www.asianacargo.com/tracking/viewTraceAirWaybill.do')
    wait = WebDriverWait(driver, 10)
    row_num = row_num
    for awb_num in list:
        # 브라우저가 나올때까지 대기 및 검색버튼 나오는지 확인 후 다음스탭 진행
        wait.until(EC.element_to_be_clickable((By.ID, 'btn_search')))
        driver.find_element_by_xpath('//*[@id="searchForm"]/div[1]/div/input[2]').send_keys(awb_num[3:])
        track_button = driver.find_element_by_xpath('//*[@id="btn_search"]')
        track_button.click()
        # 수량/중량 text나오는지 검사
        wait.until(EC.element_to_be_clickable((By.CLASS_NAME, 'panel-title')))
        try:
            pol_atd = driver.find_element_by_xpath('//*[@id="R'+awb_num+'"]/div/div/div[1]/div[2]/div/div[1]/div[3]')
            pol_atd = convert_oz_time_info(pol_atd.text,'atd')
            ws.cell(row_num, 2).value = pol_atd
        except:
            ws.cell(row_num, 2).value = 'No_data'
        try:
            pod_ata = driver.find_element_by_xpath('//*[@id="R'+awb_num+'"]/div/div/div[1]/div[2]/div/div[1]/div[4]')
            pod_ata = convert_oz_time_info(pod_ata.text,'ata')
            ws.cell(row_num, 3).value = pod_ata
        except:
            ws.cell(row_num, 3).value = 'No_data'
        driver.find_element_by_xpath('//*[@id="searchForm"]/div[1]/div/input[2]').clear()
        time.sleep(0.1)
        row_num += 1
    print('아시아나 tracking 완료')
    driver.quit()
    return row_num

def convert_ke_time_info(value):
    '''대한항공 트래킹 웹사이트에서 복사된 일자값을 통일된 양식(2020-05-20 18:00)으로 가공'''
    value = value
    month_dic = {'Jan':'01', 'Feb':'02', 'Mar':'03', 'Apr':'04', 'May':'05', 'Jun':'06', 'Jul':'07', 'Aug':'08', 'Sep':'09', 'Oct':'10', 'Nov':'11', 'Dec':'12'}
    month_info = month_dic[value[3:6]]
    date_info = value[:2]
    time_info = value[6:]
    result = '2020-'+month_info+'-'+date_info+' '+time_info
    # 크롤링된 값 공통 값으로 정리하기(ex) 24 Mar 2020 18:00 -> 2020-03-24 18:00)
    return result

def convert_oz_time_info(value,info):
    '''대한항공 트래킹 웹사이트에서 복사된 일자값을 통일된 양식(2020-05-20 18:00)으로 가공'''
    if info == 'atd':
        result = value[17:]
    elif info == 'ata':
        result = value[15:]
    # 크롤링된 값 공통 값으로 정리하기(ex) departure date 2020 18:00 -> 2020-03-24 18:00)
    return result

# 안내메시지 나가기 ('AWB 넘버를 복사하셨나요?')

origin_folder = 'D:\\'
driver_dir = r'C:\chrome_driver\chromedriver.exe'
file_subject = get_subject()
os.chdir(origin_folder)
wb = Workbook()
ws = wb.active
row_num = 1
copy_v = pyperclip.paste()
divided_values = value_make(copied_value=copy_v)
# AWB 앞3자리가 대한항공(180), 아시아나(988) AWB 만 각 LIST 에 넣기
total_list, ke_list, oz_list = verifi_awb(divided_values)
max_row = len(total_list)
for i in range(max_row):
    ws.cell(i+1, 1).value = total_list[i]
# awb별 tracking 정보 반환 및 엑셀에 저장
row_num = get_tracking_info_ke(ke_list, driver_dir, row_num)
row_num = get_tracking_info_oz(oz_list, driver_dir, row_num)
file_save(origin_folder=origin_folder, file_subject = file_subject)
print('후로그램 끝 ^0^ by 세환선임')
print('총소요시간(초) : ')
print(time.perf_counter())