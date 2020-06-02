import os
from openpyxl import Workbook
from openpyxl import load_workbook

origin_path = 'd:\\'
origin_excel_file_name = 'sample.xlsx'
os.chdir(origin_path)

wb = load_workbook(origin_excel_file_name)
ws = wb.active

# 시트의 최대 row, column값 확인
max_column = ws.max_column
max_row = ws.max_row

# Code로 시작하는 row 넘버 확인(변수 : row_start)
for i in range(1, max_row):
    if((ws.cell(row = i,column = 1).value)=='Code'):
        row_start = i

# 반복행수 확인(변수 : re_row_no)
re_row_no = max_row - row_start

# 반복칼럼수 확인(변수 : re_column_no)
re_column_no = max_column - 2

# 변환된값이 입력될 대상 wb생성
target_wb = Workbook()
target_ws = target_wb.active

target_ws.cell(row = 1, column = 1).value = '사명'
target_ws.cell(row = 1, column = 2).value = '분기'
target_ws.cell(row = 1, column = 3).value = '일시'
target_ws.cell(row = 1, column = 4).value = '코드'
target_ws.cell(row = 1, column = 5).value = '코드명'
target_ws.cell(row = 1, column = 6).value = '값'

# 최대반복행수 정의
max_re_row_no = re_row_no * re_column_no

# 타겟 워크시트 기초 칼럼명 입력(1)_사명
for i in range(1, max_re_row_no+1):
    target_ws.cell(row = i + 1, column = 1).value = ws.cell(row = 4, column = 1).value

# 타겟 워크시트 기초 칼럼명 입력(2)_분기
init_ws_column = 3
target_ws_row_no = 2
target_ws_column_no = 2
for i in range(1, re_column_no + 1):
    for j in range(1, re_row_no + 1):
        target_ws.cell(row = target_ws_row_no, column = target_ws_column_no).value = ws.cell(row = row_start - 1, column = init_ws_column).value
        target_ws_row_no += 1
    init_ws_column += 1

# 타겟 워크시트 기초 칼럼명 입력(3)_일시
init_ws_column = 3
target_ws_row_no = 2
target_ws_column_no = 3
for i in range(1, re_column_no + 1):
    for j in range(1, re_row_no + 1):
        target_ws.cell(row = target_ws_row_no, column = target_ws_column_no).value = \
            ws.cell(row = row_start, column = init_ws_column).value
        target_ws_row_no += 1
    init_ws_column += 1

# 타겟 워크시트 기초 칼럼명 입력(4)_코드
init_ws_column = 1
target_ws_row_no = 2
target_ws_column_no = 4
for i in range(1, re_column_no + 1):
    row_plus = 1
    for j in range(1, re_row_no + 1):
        target_ws.cell(row = target_ws_row_no, column = target_ws_column_no).value = \
            ws.cell(row = row_start+row_plus, column = init_ws_column).value
        target_ws_row_no += 1
        row_plus +=1
    init_ws_column = 1

# 타겟 워크시트 기초 칼럼명 입력(5)_코드명
init_ws_column = 2
target_ws_row_no = 2
target_ws_column_no = 5
for i in range(1, re_column_no + 1):
    row_plus = 1
    for j in range(1, re_row_no + 1):
        target_ws.cell(row = target_ws_row_no, column = target_ws_column_no).value = \
            ws.cell(row = row_start+row_plus, column = init_ws_column).value
        target_ws_row_no += 1
        row_plus +=1

# 타겟 워크시트 기초 칼럼명 입력(6)_값
init_ws_column = 3
target_ws_row_no = 2
target_ws_column_no = 6
for i in range(1, re_column_no + 1):
    row_plus = 1
    for j in range(1, re_row_no + 1):
        target_ws.cell(row = target_ws_row_no, column = target_ws_column_no).value = \
            ws.cell(row = row_start+row_plus, column = init_ws_column).value
        target_ws_row_no += 1
        row_plus += 1
    init_ws_column += 1

# 타겟 워크북 저장
target_wb.save(origin_path+'sample_py_result.xlsx')
