import os
from openpyxl import Workbook
from openpyxl import load_workbook
import pyautogui
from tkinter import Tk
from datetime import datetime
import pyperclip

def image_search_click(file_name, route):
    # ※주의※ 파일명 영어만 됨
    os.chdir(route)
    while True:
        location = pyautogui.locateCenterOnScreen(file_name)
        if location == None:
            continue
        else:
            pyautogui.doubleClick(location)
            pyautogui.PAUSE = 0.1
            break

def image_search_click2(file_name, route):
    # ※주의※ 파일명 영어만 됨
    os.chdir(route)
    no_result = 'no_result.png'
    while True:
        location = pyautogui.locateCenterOnScreen(file_name)
        location2 = pyautogui.locateCenterOnScreen(no_result)

        if location == None and location2 == None:
            continue
        if location != None:
            pyautogui.doubleClick(location)
            break
        else:
            pyautogui.doubleClick(location2)
            break

def file_save(origin_folder, file_subject):
    os.chdir(origin_folder)
    wb.save(file_subject)

def value_make(copied_value, column):
    """ 복사된 값의 성명, 직위등 추출"""
    name = column
    divide_values = copied_value.split()

    for i in range(len(divide_values)):
        if divide_values[i] == name:
            return divide_values[i+1]

def image_search_only(file_name, route):
    # ※주의※ 파일명 영어만 됨
    os.chdir(route)
    a=1
    message = 'not found'
    while True:
        location = pyautogui.locateCenterOnScreen(file_name)
        if location == None:
            pyautogui.PAUSE = 0.1
            print(file_name+'이미지 찾지 못함')
            a += 1
            if a>20:
                return message
            break
        else:
            print(file_name+'이미지 찾음')
            pyautogui.PAUSE = 0.1
            break

# 엑셀파일 만들기
origin_folder = 'D:\\세환\\python\\사번추출'
image_folder = 'D:\\세환\\python\\사번추출\\image'
os.chdir(origin_folder)
wb = load_workbook(''
                   '.xlsx')
ws = wb.active
init_cell_row = 1
init_cell_col = 1

x,y = pyautogui.size()
x = x/2
y = y/4*3

while True:
    current_cell = ws.cell(init_cell_row, init_cell_col)
    if current_cell.value == None:
        break
    # 저장된 엑셀객체(ws) A'i'셀 에 값(사번) 넣기
    # 이미지 더블클릭('성명/email/사번.png') 및 탭하여 오른쪽 클릭(사번입력칸)
    image_search_click(file_name='name.png', route=image_folder)
    pyautogui.press('tab')
    pyautogui.PAUSE = 0.1
    # 값(사번) 붙여넣기
    pyperclip.copy(current_cell.value)
    pyautogui.hotkey('ctrl','v')
    pyautogui.press('enter')
    pyautogui.PAUSE = 0.1
    # 값이 나왔는지 여부 보기
   # message = image_search_only('no_result.png', image_folder)
    #if message != 'not found':
     #   image_search_click(file_name='basic_info.png', route=image_folder)
    #else:
    # image_search_click2(file_name='basic_info.png', route=image_folder)
    pyautogui.doubleClick(x,y)
    pyautogui.PAUSE = 0.1
    # cntr+a 핫키 입력 하여 전체 텍스트 복사
    pyautogui.hotkey('ctrl', 'a')
    pyautogui.PAUSE = 0.1
    pyautogui.hotkey('ctrl', 'c')
    pyautogui.PAUSE = 0.1
    copied_value = Tk().clipboard_get()
    copied_name = value_make(copied_value=copied_value, column='성명')
    last_copied_name = copied_name
    ws.cell(init_cell_row, init_cell_col + 2).value = copied_name
    ws.cell(init_cell_row, init_cell_col + 3).value = copied_value
    file_subject = 'employee_no_detail_result_hq' + '.xlsx'
    file_save(origin_folder=origin_folder, file_subject=file_subject)
    init_cell_row += 1

