#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
from openpyxl import Workbook
from datetime import timedelta
from datetime import datetime
from datetime import date



def get_vessel_name(a):
    for y in range(len(a)-1):
        try:
            check_matter = a[y]+a[y+1]
        except:
            pass
        if check_matter == '::':
            vessel_name = a[y+2:]
            return vessel_name


def get_carrier_name(a):
    carrier_name = ''
    for y in range(len(a)-1):
        check_matter = a[y]+a[y+1]
        if check_matter == '++':
            index_num = y+2
            break
    while True:
        carrier_name = carrier_name+a[index_num]
        index_num += 1
        if a[index_num] == ':':
            break
    return carrier_name


def get_subject(title):
    """파일제목 마지막에 저장할 현재일시 및 파일제목 설정"""
    now = datetime.now()
    today = date.today()
    current_time = now.strftime("%H-%M-%S")
    file_name = title + str(today) + '_' + str(current_time) + '.xlsx'
    return file_name


def convert_schedule_elements(schedule_elements):
    ele_dic = {'File_Name':'', '출발항':'','도착항':'','EDI문서 전송일시':'', '서류마감일':'', '카고마감일':'', 'ETD':'',
               'ETA':'','Scheduled.T.D':'', 'Vessel Name':'','선사':''}
    for a in schedule_elements:
        if 'LOC+88' in a:
            ele_dic['출발항'] = a[7:12]
        if 'LOC+7' in a:
            ele_dic['도착항'] = a[6:11]
        if 'DTM+137' in a:
            ele_dic['EDI문서 전송일시'] = a[8:18]
        if 'DTM+222' in a:
            ele_dic['서류마감일'] = a[8:18]
        if 'DTM+180' in a:
            ele_dic['카고마감일'] = a[8:18]
        if 'DTM+133' in a:
            ele_dic['ETD'] = a[8:18]
        if 'DTM+132' in a:
            ele_dic['ETA'] = a[8:18]
        if 'DTM+189' in a:
            ele_dic['Scheduled.T.D'] = a[8:18]
        if 'TDT+20' in a:
            vsl_name = get_vessel_name(a)
            ele_dic['Vessel Name'] = vsl_name
        if 'TDT+20' in a:
            carrier_name = get_carrier_name(a)
            ele_dic['선사'] = carrier_name
        if '.edi' in a:
            ele_dic['File_Name'] = a[12:]
    return ele_dic


os.chdir(r'D:\2.운영자동화자료\해운스케쥴시스템(Sea Booking Schedule MGT)\KL-Net 검증(edi)')
file_list = os.listdir()
eace_schedule = []

each_schedules = []
# 폴더내 전체 파일을 불러옴
for i in range(len(file_list)):
    file_object  = open(file_list[i], "r")
    # 개별 파일을 오픈함
    contents = file_object.read()
    a= contents.split('UNH+')
    for y in a:
        each_schedules.append('file_name : '+file_list[i]+"'"+y)
    file_object.close()

os.chdir('D:\\')
wb = Workbook()
ws = wb.active
file_subject = get_subject('kl_net ')
ele_dic = {'File_Name':'', '출발항':'','도착항':'','EDI문서 전송일시':'', '서류마감일':'', '카고마감일':'', 'ETD':'',
               'ETA':'','Scheduled.T.D':'', 'Vessel Name':'','선사':''}

cell_r = 1
cell_c = 1

for i in ele_dic.keys():
    ws.cell(cell_r, cell_c).value = i
    cell_c += 1


cell_r = 2
cell_c = 1
schedule_list = []
for schedule in each_schedules:
    schedule_elements = schedule.split("'")
    schedule_dic = convert_schedule_elements(schedule_elements)
    for key, value in schedule_dic.items():
        ws.cell(cell_r, cell_c).value = value
        cell_c += 1
    cell_c = 1
    cell_r += 1

wb.save(file_subject)
