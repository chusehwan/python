import os
from openpyxl import Workbook
import pyperclip
import time
import tracking_function as fn
import KE
import OZ

origin_folder = 'D:\\'
driver_dir = r'C:\chrome_driver\chromedriver.exe'
file_subject = fn.get_subject()
os.chdir(origin_folder)

wb = Workbook()
ws = wb.active
ws.cell(1, 1).value = 'AWB No'
ws.cell(1, 2).value = 'POL ETD'
ws.cell(1, 3).value = 'POL ATD'
ws.cell(1, 4).value = 'POD ETA'
ws.cell(1, 5).value = 'POD ATA'
ws.cell(1, 6).value = '이상유무'
row_num = 2
copy_v = pyperclip.paste()
divided_values = fn.value_make(copied_value=copy_v)
# AWB 앞3자리가 대한항공(180), 아시아나(988) AWB 만 각 LIST 에 넣기
total_list, ke_list, oz_list = fn.verifi_awb(divided_values)
# 엑셀 결과 파일에 들어갈 최대 행수 구하기
max_row = len(total_list)
for i in range(max_row):
    ws.cell(i+2, 1).value = total_list[i]
# awb별 tracking 정보 반환 및 엑셀에 저장
row_num = KE.get_tracking_info_ke(ke_list, driver_dir, row_num, ws)
row_num = OZ.get_tracking_info_oz(oz_list, driver_dir, row_num, ws)
# 최종 결과물 파일로 저장
fn.file_save(origin_folder=origin_folder, file_subject=file_subject, wb=wb)
print('프로그램 끝 ^0^')
print('총소요시간(초) : ', round(time.perf_counter(), 2))
print('건당소요시간(초) : ', round((time.perf_counter()/max_row), 2))
